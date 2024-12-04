from django.shortcuts import render,redirect
from ramapp.forms import ReadCSV
import openpyxl
from ramapp.custom_cmd import*
from django.contrib import messages
from ramapp.get_model import get_model_name
from django.apps import apps
# Create your views here.

def auto_migration(request):
    if request.method == 'GET':
        template_name = 'ramapp/code_automation.html'
        return render(request, template_name)
    if request.method == 'POST':
        get_code_automation_request=request.POST.get('get_code_automation_request')
        if get_code_automation_request == 'DoMigrations':
            do_migrations=custom_makemigrations()
            messages.success(request, 'Migrations Has Been  Done')
        elif get_code_automation_request == 'DoMigrate':
            do_migrate=custom_migrate()
            messages.success(request, 'Migrate Has Been  Done')
        return redirect('data_import')

def data_import(request):
    if request.method == 'GET':
        all_model=get_model_name()
        form = ReadCSV()
        context = {
            'form': form,'all_model':all_model,
        }
        template_name = 'ramapp/index.html'
        return render(request, template_name, context)
    elif request.method == 'POST':
        try:
            form = ReadCSV(request.POST, request.FILES)
            if form.is_valid():
                selected_table = form.cleaned_data['Select_Table']
                upload_at = form.cleaned_data['upload_at']
                print('upload_at ',upload_at)
                print('selected_table ',selected_table)
                open_workbooks = openpyxl.load_workbook(upload_at)
                print('open_workbooks ',open_workbooks)
                dataframe = open_workbooks.active
                sheet_name = open_workbooks.sheetnames  # Get all sheet Name
                data_dict={}
                dict_list=[]
                for row in dataframe.iter_rows(min_row=1, max_row=1, values_only=True):
                    # Use the values in the row as the headers
                    headers = list(row)
                    print('headers ',headers)
                for row in range(2, dataframe.max_row):
                    value_lists = []
                    for col in dataframe.iter_cols(1, dataframe.max_column):
                        value_lists.append(col[row].value)
                    print('value lists ',value_lists) 
                    data_dict=dict(zip(headers,value_lists))  
                    dict_list.append(data_dict)
                    ModelClass = apps.get_model(app_label='testapp', model_name=selected_table)
                    my_data = ModelClass(**data_dict)
                    my_data.save()               
                                             
            context = {
                'form': form,
            }
            template_name = 'ramapp/index.html'
            print('template_name=====,',template_name)
            messages.error(request,f'SUCCESS : DATA INSERTED SUCCESSFULLY')
            return render(request, template_name, context)
        except Exception as error:
            print('error ',error)
            messages.error(request,f'ERROR : {error}')
            return redirect('data_import')
